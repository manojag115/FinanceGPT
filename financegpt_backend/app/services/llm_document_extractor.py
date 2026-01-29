"""
LLM-based document extraction service for investment holdings.

Uses instructor + LLM to extract structured data from any document format:
- PDFs (brokerage statements, account summaries)
- Excel/CSV files with non-standard formats
- Images of statements
- Word documents

This provides a flexible, format-agnostic approach that handles variations
across different financial institutions.
"""

import logging
from typing import Optional, BinaryIO
from io import BytesIO

import instructor
from openai import AsyncOpenAI
from pydantic import BaseModel, Field

from app.config import config
from app.schemas.investments import (
    InvestmentHoldingEnriched,
    AccountTaxType,
)

logger = logging.getLogger(__name__)


class ExtractedHoldingsData(BaseModel):
    """Structured data extracted from a document."""
    
    account_name: Optional[str] = Field(
        None,
        description="Account name or description (e.g., 'Individual Brokerage', 'Roth IRA')"
    )
    account_number: Optional[str] = Field(
        None,
        description="Account number if visible in document"
    )
    account_type: Optional[str] = Field(
        None,
        description="Type of account: brokerage, 401k, ira_traditional, ira_roth, etc."
    )
    institution: Optional[str] = Field(
        None,
        description="Financial institution name (e.g., 'Fidelity', 'Vanguard', 'Charles Schwab')"
    )
    statement_date: Optional[str] = Field(
        None,
        description="Statement or report date in ISO format (YYYY-MM-DD)"
    )
    holdings: list[dict] = Field(
        default_factory=list,
        description="List of investment holdings with symbol, quantity, cost_basis, etc."
    )
    
    class Config:
        json_schema_extra = {
            "example": {
                "account_name": "Individual Brokerage",
                "account_number": "Z12345678",
                "account_type": "brokerage",
                "institution": "Fidelity",
                "statement_date": "2026-01-29",
                "holdings": [
                    {
                        "symbol": "AAPL",
                        "description": "Apple Inc",
                        "quantity": 100,
                        "cost_basis": 15000.00,
                        "average_cost_basis": 150.00,
                    }
                ]
            }
        }


class LLMDocumentExtractor:
    """Extract investment holdings from documents using LLM."""
    
    def __init__(self):
        """Initialize the LLM client with instructor."""
        # Use the configured LLM from FinanceGPT settings
        self.client = instructor.from_openai(
            AsyncOpenAI(
                api_key=config.OPENAI_API_KEY,
                base_url=config.OPENAI_BASE_URL if hasattr(config, 'OPENAI_BASE_URL') else None,
            )
        )
    
    async def extract_from_text(
        self,
        text_content: str,
        account_tax_type: Optional[AccountTaxType] = None,
    ) -> dict:
        """
        Extract holdings from plain text content.
        
        Args:
            text_content: Text extracted from document
            account_tax_type: Optional tax type hint
            
        Returns:
            Dictionary with account_data and holdings
        """
        system_prompt = """You are a financial document parser. Extract investment holdings data from the provided text.

Extract the following information:
1. Account details (name, number, type, institution)
2. For each holding:
   - Symbol (ticker symbol)
   - Description (company/fund name)
   - Quantity (number of shares/units)
   - Cost Basis (total purchase cost)
   - Average Cost Basis (cost per share)
   - Current Price (if available)
   - Market Value (if available)

Important:
- Only extract data you can see in the document
- Use null/None for missing fields
- Preserve exact numeric values
- Extract ALL holdings you find
- Account type should be one of: brokerage, 401k, ira_traditional, ira_roth, 529, hsa
"""

        user_prompt = f"""Extract investment holdings from this document:

{text_content}

{f"Note: This is a {account_tax_type} account." if account_tax_type else ""}
"""

        try:
            # Use instructor to get structured output
            extracted_data = await self.client.chat.completions.create(
                model="gpt-4o",  # Use GPT-4 for better extraction accuracy
                response_model=ExtractedHoldingsData,
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": user_prompt},
                ],
                temperature=0,  # Deterministic extraction
            )
            
            # Convert to our format
            account_data = {
                "account_name": extracted_data.account_name or "Uploaded Account",
                "account_number": extracted_data.account_number or f"DOC-{hash(text_content[:100]) % 100000:05d}",
                "account_type": extracted_data.account_type or "brokerage",
                "account_tax_type": account_tax_type or "taxable",
                "institution": extracted_data.institution or "Unknown",
            }
            
            # Validate and structure holdings
            holdings = []
            for holding in extracted_data.holdings:
                # Ensure required fields
                if not holding.get("symbol"):
                    logger.warning(f"Skipping holding without symbol: {holding}")
                    continue
                
                # Build holding dict with required fields
                holding_dict = {
                    "symbol": holding["symbol"],
                    "description": holding.get("description"),
                    "quantity": float(holding.get("quantity", 0)),
                    "cost_basis": float(holding.get("cost_basis", 0)),
                    "average_cost_basis": float(holding.get("average_cost_basis", 0)),
                }
                
                # Add optional fields if present
                if "current_price" in holding and holding["current_price"]:
                    holding_dict["current_price"] = float(holding["current_price"])
                if "market_value" in holding and holding["market_value"]:
                    holding_dict["market_value"] = float(holding["market_value"])
                
                holdings.append(holding_dict)
            
            return {
                "account_data": account_data,
                "holdings": holdings,
                "extraction_method": "llm",
            }
            
        except Exception as e:
            logger.error(f"LLM extraction failed: {e}")
            raise ValueError(f"Failed to extract holdings from document: {str(e)}")
    
    async def extract_from_pdf(
        self,
        pdf_file: BinaryIO,
        account_tax_type: Optional[AccountTaxType] = None,
    ) -> dict:
        """
        Extract holdings from PDF file.
        
        Args:
            pdf_file: PDF file binary stream
            account_tax_type: Optional tax type hint
            
        Returns:
            Dictionary with account_data and holdings
        """
        try:
            import pdfplumber
        except ImportError:
            raise ImportError("pdfplumber is required for PDF parsing. Install with: pip install pdfplumber")
        
        # Extract text and tables from PDF
        text_content = []
        
        with pdfplumber.open(pdf_file) as pdf:
            # Extract text from all pages
            for page in pdf.pages:
                text_content.append(page.extract_text())
                
                # Also extract tables which might contain holdings
                tables = page.extract_tables()
                for table in tables:
                    # Convert table to readable text format
                    if table:
                        table_text = "\n".join(["\t".join(str(cell) for cell in row if cell) for row in table])
                        text_content.append(table_text)
        
        # Combine all extracted content
        full_text = "\n\n".join(filter(None, text_content))
        
        if not full_text.strip():
            raise ValueError("No text could be extracted from the PDF")
        
        # Use LLM to extract structured data
        return await self.extract_from_text(full_text, account_tax_type)
    
    async def extract_from_excel(
        self,
        excel_file: BinaryIO,
        account_tax_type: Optional[AccountTaxType] = None,
    ) -> dict:
        """
        Extract holdings from Excel file.
        
        Args:
            excel_file: Excel file binary stream
            account_tax_type: Optional tax type hint
            
        Returns:
            Dictionary with account_data and holdings
        """
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl is required for Excel parsing. Install with: pip install openpyxl")
        
        # Load workbook
        workbook = openpyxl.load_workbook(excel_file, data_only=True)
        
        # Extract all sheets as text
        text_content = []
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text_content.append(f"Sheet: {sheet_name}")
            
            # Convert sheet to text format
            for row in sheet.iter_rows(values_only=True):
                row_text = "\t".join(str(cell) if cell is not None else "" for cell in row)
                if row_text.strip():
                    text_content.append(row_text)
        
        # Combine all content
        full_text = "\n".join(text_content)
        
        if not full_text.strip():
            raise ValueError("No data could be extracted from the Excel file")
        
        # Use LLM to extract structured data
        return await self.extract_from_text(full_text, account_tax_type)
    
    async def extract_from_image(
        self,
        image_file: BinaryIO,
        account_tax_type: Optional[AccountTaxType] = None,
    ) -> dict:
        """
        Extract holdings from image file using GPT-4 Vision.
        
        Args:
            image_file: Image file binary stream
            account_tax_type: Optional tax type hint
            
        Returns:
            Dictionary with account_data and holdings
        """
        import base64
        
        # Read image and encode to base64
        image_bytes = image_file.read()
        base64_image = base64.b64encode(image_bytes).decode('utf-8')
        
        # Detect image format
        image_format = "jpeg"  # default
        if image_bytes.startswith(b'\x89PNG'):
            image_format = "png"
        elif image_bytes.startswith(b'GIF'):
            image_format = "gif"
        
        system_prompt = """You are a financial document parser. Extract investment holdings data from the provided image.

Extract the following information:
1. Account details (name, number, type, institution)
2. For each holding:
   - Symbol (ticker symbol)
   - Description (company/fund name)
   - Quantity (number of shares/units)
   - Cost Basis (total purchase cost)
   - Average Cost Basis (cost per share)
   - Current Price (if available)
   - Market Value (if available)

Important:
- Only extract data you can see in the image
- Use null/None for missing fields
- Preserve exact numeric values
- Extract ALL holdings you find
- Account type should be one of: brokerage, 401k, ira_traditional, ira_roth, 529, hsa
"""

        user_prompt = f"Extract investment holdings from this statement image."
        if account_tax_type:
            user_prompt += f"\nNote: This is a {account_tax_type} account."
        
        try:
            # Use GPT-4 Vision to extract data
            extracted_data = await self.client.chat.completions.create(
                model="gpt-4o",
                response_model=ExtractedHoldingsData,
                messages=[
                    {"role": "system", "content": system_prompt},
                    {
                        "role": "user",
                        "content": [
                            {"type": "text", "text": user_prompt},
                            {
                                "type": "image_url",
                                "image_url": {
                                    "url": f"data:image/{image_format};base64,{base64_image}"
                                }
                            }
                        ]
                    },
                ],
                temperature=0,
            )
            
            # Convert to our format (same as extract_from_text)
            account_data = {
                "account_name": extracted_data.account_name or "Uploaded Account",
                "account_number": extracted_data.account_number or f"IMG-{hash(base64_image[:100]) % 100000:05d}",
                "account_type": extracted_data.account_type or "brokerage",
                "account_tax_type": account_tax_type or "taxable",
                "institution": extracted_data.institution or "Unknown",
            }
            
            holdings = []
            for holding in extracted_data.holdings:
                if not holding.get("symbol"):
                    continue
                
                holding_dict = {
                    "symbol": holding["symbol"],
                    "description": holding.get("description"),
                    "quantity": float(holding.get("quantity", 0)),
                    "cost_basis": float(holding.get("cost_basis", 0)),
                    "average_cost_basis": float(holding.get("average_cost_basis", 0)),
                }
                
                if "current_price" in holding and holding["current_price"]:
                    holding_dict["current_price"] = float(holding["current_price"])
                if "market_value" in holding and holding["market_value"]:
                    holding_dict["market_value"] = float(holding["market_value"])
                
                holdings.append(holding_dict)
            
            return {
                "account_data": account_data,
                "holdings": holdings,
                "extraction_method": "llm_vision",
            }
            
        except Exception as e:
            logger.error(f"Vision extraction failed: {e}")
            raise ValueError(f"Failed to extract holdings from image: {str(e)}")
